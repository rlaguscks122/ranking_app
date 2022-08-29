import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet('rank')

ws['A1'] = 'name'
ws['B1'] = 'rank'

ws['D1'] = 'name'
ws['E1'] = 'rank'

wb.save(r'C:\python-codes\ranking-test\rankingdate\rank_data.xlsx')

