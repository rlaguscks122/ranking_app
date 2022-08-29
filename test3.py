from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

row = 2
while True:
    fpath = r'rank_data.xlsx'
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    unisent = webdriver.Chrome("/usr/bin/chromedriver")
    url="https://sjnim.com/"
    unisent.get(url)
    unisent.implicitly_wait(5)
    unisent.maximize_window()

    move_page = unisent.find_element(By.CSS_SELECTOR, "#store_search_show")
    move_page.click()

    move_kakao = unisent.find_element(By.CSS_SELECTOR, "#sns_login > div > a.sns-icon.social_link.sns-kakao")
    move_kakao.click()

    print(unisent.window_handles)

    unisent.switch_to.window(unisent.window_handles[-1])

    id = unisent.find_element(By.CSS_SELECTOR, '#id_email_2')
    id.send_keys("girina79@naver.com")

    pw = unisent.find_element(By.CSS_SELECTOR, '#id_password_3')
    pw.send_keys("kkk123")

    login_btn = unisent.find_element(By.XPATH, '//*[@id="login-form"]/fieldset/div[8]/button[1]')
    login_btn.click()

    unisent.switch_to.window(unisent.window_handles[0])

    close_pop = unisent.find_element(By.CSS_SELECTOR, '#hd_pops_6 > div.hd_pops_footer > button.hd_pops_close.hd_pops_6')
    close_pop.click()

    unisent.switch_to.window(unisent.window_handles[0])

    name = unisent.find_element(By.CSS_SELECTOR, '#storename')
    name.send_keys("겸조")

    keyword = unisent.find_element(By.CSS_SELECTOR, '#keyword')
    keyword.send_keys("메이킹미")

    search_button = unisent.find_element(By.CSS_SELECTOR, '#keyword_search')
    search_button.click()

    item_one = unisent.find_element(By.CSS_SELECTOR, '#product_list > section:nth-child(1) > div > div')

    info_one = item_one.find_element(By.CSS_SELECTOR, '#product_list > section:nth-child(1) > div > div > div.product_info > h6').text
    rank_one = item_one.find_element(By.CSS_SELECTOR, "#product_list > section:nth-child(1) > div > div > div.product_info > h4").text
    print(info_one, rank_one)
    ws[f'A{row}'] = info_one
    ws[f'B{row}'] = rank_one
    row = row + 1
    
    item_two = unisent.find_element(By.CSS_SELECTOR, '#product_list > section:nth-child(2) > div > div')

    info_two = item_two.find_element(By.XPATH, '//*[@id="product_list"]/section[2]/div/div/div[2]/h6').text
    rank_two = item_two.find_element(By.CSS_SELECTOR, '#product_list > section:nth-child(2) > div > div > div.product_info > h4').text
    print(info_two, rank_two)
    ws[f'D{row}'] = info_two
    ws[f'E{row}'] = rank_two
    row = row + 1

    wb.save(fpath)


    unisent.quit()
    print("2분마다 가져올것")
    time.sleep(120)


