import openpyxl

fpath = r'C:\python-codes\ranking-test\rank_data.xlsx'

wb = openpyxl.load_workbook(fpath)

ws = wb['rank']

ws['A2'] = 23
ws['B2'] = 'ddd'

wb.save(fpath)